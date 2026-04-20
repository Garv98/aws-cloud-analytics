import csv
import json
import os
import re
import math
import random
from collections import defaultdict
from datetime import datetime, timezone
from io import BytesIO, StringIO
from urllib.parse import unquote_plus
from statistics import median as stats_median

import boto3
from botocore.exceptions import ClientError

# Try to import Excel libraries
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

try:
    import xlrd
    XLS_SUPPORT = True
except ImportError:
    XLS_SUPPORT = False


dynamodb = boto3.resource("dynamodb")
s3_client = boto3.client("s3")


def _now_iso():
    return datetime.now(timezone.utc).isoformat()


def _required_env(name: str) -> str:
    value = os.environ.get(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def _job_id_from_key(key: str) -> str:
    filename = key.split("/")[-1]
    base_name = filename.rsplit(".", 1)[0]
    return base_name


def _detect_file_type(key: str, content: bytes) -> str:
    """Detect file type from extension and content."""
    lower_key = key.lower()
    
    if lower_key.endswith(".xlsx"):
        return "xlsx"
    elif lower_key.endswith(".xls"):
        return "xls"
    elif lower_key.endswith(".csv"):
        return "csv"
    
    # Try to detect from content
    if content[:4] == b'PK\x03\x04':  # ZIP signature (xlsx)
        return "xlsx"
    elif content[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':  # OLE signature (xls)
        return "xls"
    
    return "csv"


def _parse_excel_xlsx(content: bytes):
    """Parse XLSX file using openpyxl."""
    if not EXCEL_SUPPORT:
        raise ValueError("XLSX support not available")
    
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = wb.active
    
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel file is empty")
    
    headers = [str(h) if h else f"Column_{i}" for i, h in enumerate(rows[0])]
    data = []
    
    for row in rows[1:]:
        if any(cell is not None for cell in row):
            data.append(dict(zip(headers, row)))
    
    wb.close()
    return headers, data


def _parse_excel_xls(content: bytes):
    """Parse XLS file using xlrd."""
    if not XLS_SUPPORT:
        raise ValueError("XLS support not available")
    
    wb = xlrd.open_workbook(file_contents=content)
    sheet = wb.sheet_by_index(0)
    
    if sheet.nrows == 0:
        raise ValueError("Excel file is empty")
    
    headers = [str(sheet.cell_value(0, i)) or f"Column_{i}" for i in range(sheet.ncols)]
    data = []
    
    for row_idx in range(1, sheet.nrows):
        row_data = {}
        for col_idx, header in enumerate(headers):
            cell = sheet.cell(row_idx, col_idx)
            value = cell.value
            
            # Handle Excel date format
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    value = xlrd.xldate_as_datetime(value, wb.datemode).isoformat()
                except:
                    pass
            
            row_data[header] = value
        data.append(row_data)
    
    return headers, data


def _parse_csv(content: str):
    """Parse CSV content."""
    reader = csv.DictReader(StringIO(content))
    headers = reader.fieldnames or []
    data = list(reader)
    return headers, data


def _detect_column_type(values):
    """Detect the data type of a column."""
    non_null_values = [v for v in values if v is not None and str(v).strip() != ""]
    
    if not non_null_values:
        return "empty"
    
    # Check for numeric
    numeric_count = 0
    for v in non_null_values:
        try:
            float(str(v).replace(",", ""))
            numeric_count += 1
        except (ValueError, TypeError):
            pass
    
    if numeric_count / len(non_null_values) > 0.8:
        return "numeric"
    
    # Check for datetime
    date_patterns = [
        r'\d{4}-\d{2}-\d{2}',
        r'\d{2}/\d{2}/\d{4}',
        r'\d{2}-\d{2}-\d{4}',
    ]
    date_count = 0
    for v in non_null_values:
        v_str = str(v)
        if any(re.search(p, v_str) for p in date_patterns):
            date_count += 1
    
    if date_count / len(non_null_values) > 0.8:
        return "datetime"
    
    # Check for categorical (limited unique values)
    unique_ratio = len(set(str(v) for v in non_null_values)) / len(non_null_values)
    if unique_ratio < 0.5:
        return "categorical"
    
    return "text"


def _percentile(data, p):
    """Pure Python percentile calculation."""
    sorted_data = sorted(data)
    n = len(sorted_data)
    k = (n - 1) * p / 100
    f = int(k)
    c = f + 1 if f + 1 < n else f
    return sorted_data[f] + (k - f) * (sorted_data[c] - sorted_data[f]) if f != c else sorted_data[f]


def _compute_numeric_stats(values):
    """Compute comprehensive statistics for numeric column with advanced analytics."""
    clean_values = []
    for v in values:
        if v is not None and str(v).strip() != "":
            try:
                clean_values.append(float(str(v).replace(",", "")))
            except (ValueError, TypeError):
                pass
    
    if not clean_values:
        return None
    
    n = len(clean_values)
    sorted_vals = sorted(clean_values)
    
    # Basic stats - pure Python (fast)
    total = sum(clean_values)
    mean_val = total / n
    min_val = sorted_vals[0]
    max_val = sorted_vals[-1]
    median_val = sorted_vals[n // 2] if n % 2 else (sorted_vals[n // 2 - 1] + sorted_vals[n // 2]) / 2
    
    # Variance and std
    variance = sum((x - mean_val) ** 2 for x in clean_values) / n
    std_val = math.sqrt(variance)
    
    stats_dict = {
        "count": n,
        "sum": round(total, 2),
        "mean": round(mean_val, 2),
        "median": round(median_val, 2),
        "std": round(std_val, 2),
        "variance": round(variance, 2),
        "min": round(min_val, 2),
        "max": round(max_val, 2),
    }
    
    # Quartiles - pure Python
    q1 = _percentile(clean_values, 25)
    q2 = _percentile(clean_values, 50)
    q3 = _percentile(clean_values, 75)
    iqr = q3 - q1
    
    stats_dict["q1"] = round(q1, 2)
    stats_dict["q2"] = round(q2, 2)
    stats_dict["q3"] = round(q3, 2)
    stats_dict["iqr"] = round(iqr, 2)
    
    # Skewness (simplified formula)
    if n > 2 and std_val > 0:
        skewness = sum(((x - mean_val) / std_val) ** 3 for x in clean_values) / n
        stats_dict["skewness"] = round(skewness, 3)
    else:
        stats_dict["skewness"] = 0
    
    # Outliers (IQR method)
    lower_bound = q1 - 1.5 * iqr
    upper_bound = q3 + 1.5 * iqr
    outliers = [v for v in clean_values if v < lower_bound or v > upper_bound]
    stats_dict["outlierCount"] = len(outliers)
    stats_dict["outlierPercent"] = round(len(outliers) / n * 100, 1)
    
    # Advanced analytics
    # Z-score outliers
    z_outliers = _z_score_outliers(clean_values)
    stats_dict["zScoreOutliers"] = z_outliers[:10]  # Top 10
    stats_dict["zScoreOutlierCount"] = len(z_outliers)
    
    # Modified Z-score outliers
    mz_outliers = _modified_z_score_outliers(clean_values)
    stats_dict["modifiedZScoreOutliers"] = mz_outliers[:10]  # Top 10
    stats_dict["modifiedZScoreOutlierCount"] = len(mz_outliers)
    
    # Normality test
    normality = _normality_test(clean_values)
    stats_dict["normalityTest"] = normality
    
    # Pattern detection
    patterns = _detect_patterns(clean_values)
    stats_dict["patterns"] = patterns
    
    # Clustering (if enough data)
    if n >= 10 and len(set(clean_values)) >= 3:
        clustering = _kmeans_clustering(clean_values, k=min(3, len(set(clean_values))))
        stats_dict["clustering"] = clustering
    
    return stats_dict


def _compute_categorical_stats(values):
    """Compute statistics for categorical column."""
    clean_values = [str(v).strip() for v in values if v is not None and str(v).strip() != ""]
    
    if not clean_values:
        return None
    
    freq = defaultdict(int)
    for v in clean_values:
        freq[v] += 1
    
    sorted_freq = sorted(freq.items(), key=lambda x: x[1], reverse=True)
    
    return {
        "count": len(clean_values),
        "uniqueCount": len(freq),
        "mode": sorted_freq[0][0] if sorted_freq else None,
        "modeCount": sorted_freq[0][1] if sorted_freq else 0,
        "topValues": [{"value": k, "count": v, "percent": round(v / len(clean_values) * 100, 1)} 
                      for k, v in sorted_freq[:10]],
    }


def _compute_datetime_stats(values):
    """Compute statistics for datetime column."""
    dates = []
    for v in values:
        if v is None or str(v).strip() == "":
            continue
        try:
            if isinstance(v, datetime):
                dates.append(v)
            else:
                # Try parsing common formats
                v_str = str(v)
                for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"]:
                    try:
                        dates.append(datetime.strptime(v_str[:len(fmt)+2], fmt))
                        break
                    except:
                        pass
        except:
            pass
    
    if not dates:
        return None
    
    dates.sort()
    return {
        "count": len(dates),
        "earliest": dates[0].isoformat(),
        "latest": dates[-1].isoformat(),
        "range_days": (dates[-1] - dates[0]).days,
    }


def _pearson_correlation(x, y):
    """Pure Python Pearson correlation coefficient."""
    n = len(x)
    if n < 3:
        return None
    
    mean_x = sum(x) / n
    mean_y = sum(y) / n
    
    num = sum((x[i] - mean_x) * (y[i] - mean_y) for i in range(n))
    den_x = math.sqrt(sum((xi - mean_x) ** 2 for xi in x))
    den_y = math.sqrt(sum((yi - mean_y) ** 2 for yi in y))
    
    if den_x * den_y == 0:
        return None
    
    return num / (den_x * den_y)


def _spearman_correlation(x, y):
    """Pure Python Spearman rank correlation coefficient."""
    n = len(x)
    if n < 3:
        return None
    
    # Rank the data
    def rank_data(data):
        sorted_data = sorted((val, idx) for idx, val in enumerate(data))
        ranks = [0] * n
        for rank, (_, idx) in enumerate(sorted_data):
            ranks[idx] = rank + 1
        return ranks
    
    rank_x = rank_data(x)
    rank_y = rank_data(y)
    
    return _pearson_correlation(rank_x, rank_y)


def _z_score_outliers(values, threshold=3.0):
    """Detect outliers using Z-score method."""
    if len(values) < 3:
        return []
    
    mean_val = sum(values) / len(values)
    std_val = math.sqrt(sum((x - mean_val) ** 2 for x in values) / len(values))
    
    if std_val == 0:
        return []
    
    outliers = []
    for i, val in enumerate(values):
        z_score = abs((val - mean_val) / std_val)
        if z_score > threshold:
            outliers.append({
                "index": i,
                "value": val,
                "z_score": round(z_score, 3)
            })
    
    return outliers


def _modified_z_score_outliers(values, threshold=3.5):
    """Detect outliers using Modified Z-score method (MAD-based)."""
    if len(values) < 3:
        return []
    
    median_val = stats_median(values)
    mad = stats_median([abs(x - median_val) for x in values])
    
    if mad == 0:
        return []
    
    outliers = []
    for i, val in enumerate(values):
        modified_z_score = 0.6745 * (val - median_val) / mad
        if abs(modified_z_score) > threshold:
            outliers.append({
                "index": i,
                "value": val,
                "modified_z_score": round(modified_z_score, 3)
            })
    
    return outliers


def _normality_test(values):
    """Perform Shapiro-Wilk-like normality test approximation."""
    if len(values) < 3:
        return {"is_normal": False, "p_value": None, "method": "insufficient_data"}
    
    n = len(values)
    mean_val = sum(values) / n
    std_val = math.sqrt(sum((x - mean_val) ** 2 for x in values) / n)
    
    if std_val == 0:
        return {"is_normal": False, "p_value": None, "method": "zero_variance"}
    
    # Calculate skewness and kurtosis
    skewness = sum(((x - mean_val) / std_val) ** 3 for x in values) / n
    kurtosis = sum(((x - mean_val) / std_val) ** 4 for x in values) / n - 3
    
    # Approximate normality based on skewness and kurtosis
    is_normal = abs(skewness) < 0.5 and abs(kurtosis) < 1.0
    
    # Approximate p-value based on deviation from normal
    deviation = math.sqrt(skewness ** 2 + kurtosis ** 2)
    p_value = max(0.01, min(0.99, 1 - deviation / 2))
    
    return {
        "is_normal": is_normal,
        "p_value": round(p_value, 4),
        "skewness": round(skewness, 3),
        "kurtosis": round(kurtosis, 3),
        "method": "skewness_kurtosis"
    }


def _detect_patterns(values):
    """Detect common patterns in data."""
    patterns = []
    
    if len(values) < 3:
        return patterns
    
    # Check for monotonic trend
    is_increasing = all(values[i] <= values[i+1] for i in range(len(values)-1))
    is_decreasing = all(values[i] >= values[i+1] for i in range(len(values)-1))
    
    if is_increasing:
        patterns.append({"type": "monotonic_increasing", "strength": "strong"})
    elif is_decreasing:
        patterns.append({"type": "monotonic_decreasing", "strength": "strong"})
    
    # Check for periodic pattern (simple check)
    if len(values) >= 6:
        first_half = values[:len(values)//2]
        second_half = values[len(values)//2:]
        if len(first_half) == len(second_half):
            correlation = _pearson_correlation(first_half, second_half)
            if correlation and correlation > 0.8:
                patterns.append({"type": "periodic", "correlation": round(correlation, 3)})
    
    # Check for constant values
    if len(set(values)) == 1:
        patterns.append({"type": "constant", "value": values[0]})
    
    # Check for exponential growth
    if len(values) >= 3 and all(v > 0 for v in values):
        log_values = [math.log(v) for v in values]
        log_corr = _pearson_correlation(list(range(len(values))), log_values)
        if log_corr and log_corr > 0.95:
            patterns.append({"type": "exponential_growth", "correlation": round(log_corr, 3)})
    
    return patterns


def _kmeans_clustering(values, k=3, max_iterations=10):
    """Simple K-means clustering implementation."""
    if len(values) < k or len(set(values)) < k:
        return None
    
    # Initialize centroids randomly
    centroids = random.sample(values, k)
    
    for _ in range(max_iterations):
        # Assign points to nearest centroid
        clusters = [[] for _ in range(k)]
        for val in values:
            distances = [abs(val - c) for c in centroids]
            nearest = distances.index(min(distances))
            clusters[nearest].append(val)
        
        # Update centroids
        new_centroids = []
        for cluster in clusters:
            if cluster:
                new_centroids.append(sum(cluster) / len(cluster))
            else:
                new_centroids.append(centroids[clusters.index(cluster)])
        
        if new_centroids == centroids:
            break
        centroids = new_centroids
    
    # Calculate cluster statistics
    cluster_stats = []
    for i, cluster in enumerate(clusters):
        if cluster:
            cluster_stats.append({
                "cluster_id": i,
                "centroid": round(centroids[i], 2),
                "size": len(cluster),
                "min": round(min(cluster), 2),
                "max": round(max(cluster), 2),
                "mean": round(sum(cluster) / len(cluster), 2)
            })
    
    return {
        "k": k,
        "clusters": cluster_stats,
        "total_points": len(values)
    }


def _compute_correlations(data, numeric_columns):
    """Compute correlation matrix for numeric columns with multiple methods."""
    if len(numeric_columns) < 2:
        return None
    
    # Build value lists for each column
    arrays = {}
    for col in numeric_columns:
        values = []
        for row in data:
            try:
                val = float(str(row.get(col, "")).replace(",", ""))
                values.append(val)
            except:
                values.append(None)
        arrays[col] = values
    
    # Compute correlations with multiple methods
    correlations = []
    for i, col1 in enumerate(numeric_columns):
        for col2 in numeric_columns[i+1:]:
            # Get paired non-null values
            x, y = [], []
            for j in range(len(data)):
                v1, v2 = arrays[col1][j], arrays[col2][j]
                if v1 is not None and v2 is not None:
                    x.append(v1)
                    y.append(v2)
            
            if len(x) > 2:
                pearson_corr = _pearson_correlation(x, y)
                spearman_corr = _spearman_correlation(x, y)
                
                if pearson_corr is not None:
                    correlations.append({
                        "column1": col1,
                        "column2": col2,
                        "pearson": round(pearson_corr, 3),
                        "spearman": round(spearman_corr, 3) if spearman_corr else None,
                        "method": "pearson"
                    })
    
    # Sort by absolute Pearson correlation
    correlations.sort(key=lambda x: abs(x["pearson"]), reverse=True)
    return correlations[:30]  # Top 30 correlations


def _compute_data_quality(data, headers):
    """Compute data quality metrics."""
    total_cells = len(data) * len(headers)
    missing_cells = 0
    
    for row in data:
        for header in headers:
            value = row.get(header)
            if value is None or str(value).strip() == "":
                missing_cells += 1
    
    completeness = round((total_cells - missing_cells) / total_cells * 100, 1) if total_cells > 0 else 0
    
    return {
        "totalRows": len(data),
        "totalColumns": len(headers),
        "totalCells": total_cells,
        "missingCells": missing_cells,
        "completeness": completeness,
        "qualityScore": min(100, completeness + 10) if completeness > 50 else completeness
    }


def _generate_chart_data(data, column_stats):
    """Generate data formatted for chart visualization."""
    charts = []
    
    for col_name, col_info in column_stats.items():
        col_type = col_info.get("type")
        stats = col_info.get("stats")
        
        if not stats:
            continue
        
        if col_type == "categorical":
            top_values = stats.get("topValues", [])
            if top_values:
                charts.append({
                    "column": col_name,
                    "chartType": "bar",
                    "title": f"Distribution: {col_name}",
                    "data": {
                        "labels": [v["value"][:20] for v in top_values[:8]],
                        "values": [v["count"] for v in top_values[:8]],
                        "colors": ["#6366f1", "#8b5cf6", "#a855f7", "#d946ef", "#ec4899", "#f43f5e", "#f97316", "#eab308"]
                    }
                })
        
        elif col_type == "numeric":
            # Create histogram data using pure Python
            values = []
            for row in data:
                try:
                    values.append(float(str(row.get(col_name, "")).replace(",", "")))
                except:
                    pass
            
            if values:
                # Pure Python histogram
                min_val = min(values)
                max_val = max(values)
                bin_count = 10
                bin_width = (max_val - min_val) / bin_count if max_val > min_val else 1
                
                hist = [0] * bin_count
                bin_edges = [min_val + i * bin_width for i in range(bin_count + 1)]
                
                for v in values:
                    bin_idx = min(int((v - min_val) / bin_width), bin_count - 1) if bin_width > 0 else 0
                    hist[bin_idx] += 1
                
                charts.append({
                    "column": col_name,
                    "chartType": "histogram",
                    "title": f"Distribution: {col_name}",
                    "data": {
                        "labels": [f"{round(bin_edges[i], 1)}-{round(bin_edges[i+1], 1)}" for i in range(len(hist))],
                        "values": hist,
                        "colors": ["#6366f1"] * len(hist)
                    }
                })
                
                # Box plot data
                charts.append({
                    "column": col_name,
                    "chartType": "boxplot",
                    "title": f"Box Plot: {col_name}",
                    "data": {
                        "min": stats.get("min"),
                        "q1": stats.get("q1"),
                        "median": stats.get("median"),
                        "q3": stats.get("q3"),
                        "max": stats.get("max"),
                        "outliers": stats.get("outlierCount", 0)
                    }
                })
    
    return charts


def _generate_insights(overall_stats, column_stats, correlations, data_quality):
    """Generate advanced template-based insights from the data analysis."""
    insights = []
    
    # Data overview
    insights.append(f"Analyzed {data_quality['totalRows']:,} rows across {data_quality['totalColumns']} columns with {data_quality['completeness']}% data completeness.")
    
    # Column type breakdown
    type_counts = defaultdict(int)
    for col_info in column_stats.values():
        type_counts[col_info.get("type", "unknown")] += 1
    
    type_desc = ", ".join([f"{count} {t}" for t, count in type_counts.items()])
    insights.append(f"Column types detected: {type_desc}.")
    
    # Advanced numeric insights
    for col_name, col_info in column_stats.items():
        if col_info.get("type") == "numeric":
            stats = col_info.get("stats", {})
            
            # Outlier detection insights
            if stats.get("zScoreOutlierCount", 0) > 0:
                insights.append(f"{col_name} contains {stats['zScoreOutlierCount']} statistical outliers detected via Z-score method.")
            
            if stats.get("modifiedZScoreOutlierCount", 0) > 0:
                insights.append(f"{col_name} has {stats['modifiedZScoreOutlierCount']} outliers using Modified Z-score (MAD-based) detection.")
            
            # Normality test insights
            normality = stats.get("normalityTest", {})
            if normality.get("is_normal"):
                insights.append(f"{col_name} follows a normal distribution pattern (p-value: {normality.get('p_value', 0):.3f}).")
            elif normality.get("p_value"):
                insights.append(f"{col_name} does not follow normal distribution (p-value: {normality.get('p_value', 0):.3f}, skewness: {normality.get('skewness', 0):.2f}).")
            
            # Pattern detection insights
            patterns = stats.get("patterns", [])
            for pattern in patterns:
                if pattern["type"] == "monotonic_increasing":
                    insights.append(f"{col_name} shows a strong increasing trend pattern.")
                elif pattern["type"] == "monotonic_decreasing":
                    insights.append(f"{col_name} shows a strong decreasing trend pattern.")
                elif pattern["type"] == "periodic":
                    insights.append(f"{col_name} exhibits periodic behavior with correlation {pattern.get('correlation', 0):.3f}.")
                elif pattern["type"] == "exponential_growth":
                    insights.append(f"{col_name} demonstrates exponential growth pattern (correlation: {pattern.get('correlation', 0):.3f}).")
                elif pattern["type"] == "constant":
                    insights.append(f"{col_name} has constant values - consider reviewing data quality.")
            
            # Clustering insights
            clustering = stats.get("clustering")
            if clustering:
                insights.append(f"{col_name} can be grouped into {clustering['k']} natural clusters based on K-means analysis.")
            
            # Traditional outlier insights
            if stats.get("outlierPercent", 0) > 10:
                insights.append(f"{col_name} has {stats['outlierPercent']}% outliers via IQR method - consider investigating unusual values.")
            
            skew = stats.get("skewness", 0)
            if abs(skew) > 1:
                direction = "right (positive)" if skew > 0 else "left (negative)"
                insights.append(f"{col_name} shows {direction} skew ({skew:.2f}) - data is not normally distributed.")
    
    # Advanced correlation insights
    if correlations:
        strong_corr = [c for c in correlations if abs(c.get("pearson", c.get("correlation", 0))) > 0.7]
        for corr in strong_corr[:3]:
            pearson_val = corr.get("pearson", corr.get("correlation", 0))
            spearman_val = corr.get("spearman")
            direction = "positive" if pearson_val > 0 else "negative"
            insight = f"Strong {direction} Pearson correlation ({pearson_val}) between {corr['column1']} and {corr['column2']}."
            if spearman_val:
                insight += f" Spearman rank correlation: {spearman_val}."
            insights.append(insight)
    
    # Categorical insights
    for col_name, col_info in column_stats.items():
        if col_info.get("type") == "categorical":
            stats = col_info.get("stats", {})
            if stats.get("uniqueCount", 0) <= 3:
                insights.append(f"{col_name} has only {stats['uniqueCount']} unique values - might be a good candidate for grouping or encoding.")
            
            top_values = stats.get("topValues", [])
            if top_values and top_values[0].get("percent", 0) > 50:
                insights.append(f"{col_name} is dominated by '{top_values[0]['value']}' ({top_values[0]['percent']}% of data) - potential class imbalance.")
    
    # Data quality insights
    if data_quality.get("completeness", 100) < 90:
        insights.append(f"Data quality alert: {data_quality['missingCells']:,} missing values detected ({100 - data_quality['completeness']}% incomplete).")
    elif data_quality.get("completeness", 100) >= 95:
        insights.append(f"Excellent data quality with {data_quality['completeness']}% completeness.")
    
    return insights[:15]


def _generate_ai_insights_with_bedrock(analytics_summary: dict) -> list:
    """Generate AI-powered insights using Amazon Bedrock."""
    bedrock_enabled = os.environ.get("ENABLE_BEDROCK", "false").lower() == "true"
    
    if not bedrock_enabled:
        return None
    
    try:
        bedrock = boto3.client("bedrock-runtime")
        model_id = os.environ.get("BEDROCK_MODEL_ID", "anthropic.claude-3-haiku-20240307-v1:0")
        
        # Build prompt with data summary
        prompt = f"""Analyze this dataset and provide 5-7 key business insights. Be specific and actionable.

Dataset Summary:
- Rows: {analytics_summary.get('totalRows', 0):,}
- Columns: {analytics_summary.get('totalColumns', 0)}
- Data completeness: {analytics_summary.get('completeness', 0):.1f}%

Column Analysis:
{json.dumps(analytics_summary.get('columnSummary', {}), indent=2)}

Correlations Found:
{json.dumps(analytics_summary.get('topCorrelations', []), indent=2)}

Provide insights as a JSON array of strings. Each insight should:
1. Be specific to this data
2. Suggest actionable next steps where relevant
3. Be professional and concise

Example format: ["Sales trend shows...", "Missing data in..."]

Return ONLY the JSON array, no other text."""

        # Call Bedrock
        if "anthropic" in model_id.lower():
            # Claude format
            body = json.dumps({
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 1024,
                "messages": [{"role": "user", "content": prompt}]
            })
        else:
            # Titan or other models
            body = json.dumps({
                "inputText": prompt,
                "textGenerationConfig": {
                    "maxTokenCount": 1024,
                    "temperature": 0.7
                }
            })
        
        response = bedrock.invoke_model(
            modelId=model_id,
            body=body,
            contentType="application/json",
            accept="application/json"
        )
        
        response_body = json.loads(response["body"].read())
        
        # Extract text based on model
        if "anthropic" in model_id.lower():
            ai_text = response_body.get("content", [{}])[0].get("text", "[]")
        else:
            ai_text = response_body.get("results", [{}])[0].get("outputText", "[]")
        
        # Parse JSON array from response
        ai_insights = json.loads(ai_text)
        if isinstance(ai_insights, list):
            return ai_insights[:7]
        
    except Exception as e:
        print(f"Bedrock AI insight generation failed: {e}")
    
    return None


def _process_file(content: bytes, file_type: str):
    """Process file and return comprehensive analytics."""
    
    # Parse file based on type
    if file_type == "xlsx":
        headers, data = _parse_excel_xlsx(content)
    elif file_type == "xls":
        headers, data = _parse_excel_xls(content)
    else:
        text_content = content.decode("utf-8")
        headers, data = _parse_csv(text_content)
    
    if not data:
        raise ValueError("File contains no data rows")
    
    # Analyze each column
    column_stats = {}
    numeric_columns = []
    
    for header in headers:
        values = [row.get(header) for row in data]
        col_type = _detect_column_type(values)
        
        col_info = {"type": col_type, "stats": None}
        
        if col_type == "numeric":
            col_info["stats"] = _compute_numeric_stats(values)
            numeric_columns.append(header)
        elif col_type == "categorical":
            col_info["stats"] = _compute_categorical_stats(values)
        elif col_type == "datetime":
            col_info["stats"] = _compute_datetime_stats(values)
        
        column_stats[header] = col_info
    
    # Compute overall statistics
    overall_stats = {
        "totalRows": len(data),
        "totalColumns": len(headers),
        "numericColumns": len(numeric_columns),
        "categoricalColumns": sum(1 for c in column_stats.values() if c["type"] == "categorical"),
        "datetimeColumns": sum(1 for c in column_stats.values() if c["type"] == "datetime"),
    }
    
    # Compute correlations
    correlations = _compute_correlations(data, numeric_columns)
    
    # Compute data quality
    data_quality = _compute_data_quality(data, headers)
    
    # Generate chart data
    charts = _generate_chart_data(data, column_stats)
    
    # Generate template-based insights first
    template_insights = _generate_insights(overall_stats, column_stats, correlations, data_quality)
    
    # Try AI-powered insights with Bedrock
    ai_summary = {
        "totalRows": len(data),
        "totalColumns": len(headers),
        "completeness": data_quality.get("completeness", 0),
        "columnSummary": {
            name: {
                "type": info["type"],
                "stats": {k: v for k, v in (info.get("stats") or {}).items() 
                         if k in ["mean", "median", "min", "max", "uniqueCount", "mode"]}
            }
            for name, info in list(column_stats.items())[:10]  # Limit to 10 columns
        },
        "topCorrelations": correlations[:5] if correlations else []
    }
    
    ai_insights = _generate_ai_insights_with_bedrock(ai_summary)
    insights = ai_insights if ai_insights else template_insights
    
    return {
        "overview": overall_stats,
        "dataQuality": data_quality,
        "columns": column_stats,
        "correlations": correlations,
        "charts": charts,
        "insights": insights,
        "aiPowered": ai_insights is not None,
    }


def _update_job_status(table, job_id: str, status: str, **kwargs):
    expressions = ["#s = :status", "updatedAt = :updatedAt"]
    names = {"#s": "status"}
    values = {":status": status, ":updatedAt": _now_iso()}

    for key, value in kwargs.items():
        expressions.append(f"{key} = :{key}")
        values[f":{key}"] = value

    table.update_item(
        Key={"jobId": job_id},
        UpdateExpression="SET " + ", ".join(expressions),
        ExpressionAttributeNames=names,
        ExpressionAttributeValues=values,
    )


def _process_record(record, table, report_bucket: str):
    source_bucket = record["s3"]["bucket"]["name"]
    source_key = unquote_plus(record["s3"]["object"]["key"])
    job_id = _job_id_from_key(source_key)

    _update_job_status(table, job_id, "PROCESSING")

    # Download file
    obj = s3_client.get_object(Bucket=source_bucket, Key=source_key)
    content = obj["Body"].read()
    
    # Detect file type
    file_type = _detect_file_type(source_key, content)
    
    # Process file
    analytics = _process_file(content, file_type)
    
    # Build result payload
    result_payload = {
        "jobId": job_id,
        "generatedAt": _now_iso(),
        "fileType": file_type,
        "analytics": analytics,
        "summary": "\n".join(analytics.get("insights", [])),
    }

    result_key = f"results/{job_id}.json"
    s3_client.put_object(
        Bucket=report_bucket,
        Key=result_key,
        Body=json.dumps(result_payload, default=str).encode("utf-8"),
        ContentType="application/json",
    )

    summary_text = "\n".join(analytics.get("insights", [])[:3])
    _update_job_status(table, job_id, "COMPLETED", resultKey=result_key, summary=summary_text)


def lambda_handler(event, context):
    table_name = _required_env("TABLE_NAME")
    report_bucket = _required_env("REPORT_BUCKET")

    table = dynamodb.Table(table_name)

    for record in event.get("Records", []):
        try:
            _process_record(record, table, report_bucket)
        except Exception as exc:
            key = unquote_plus(record["s3"]["object"]["key"])
            job_id = _job_id_from_key(key)
            try:
                _update_job_status(table, job_id, "FAILED", errorMessage=str(exc))
            except ClientError:
                pass

    return {"statusCode": 200, "body": json.dumps({"ok": True})}
